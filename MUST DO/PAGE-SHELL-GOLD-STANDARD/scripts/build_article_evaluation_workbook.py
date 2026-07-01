#!/usr/bin/env python3
"""Generate the Article Evaluation Excel workbook template for API/station output."""

from __future__ import annotations

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit("pip install openpyxl") from exc

OUT = Path(r"D:\GitHub\faiththruphysics-site-data\APIs\revolution-of-truth-pipeline\templates\Article_Evaluation_Workbook.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1A1A1A")
HEADER_FONT = Font(color="D4AF37", bold=True)
LABEL_FONT = Font(bold=True)


def header_row(ws, row: int, values: list[str]) -> None:
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def autosize(ws, max_col: int) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 22


def build() -> None:
    wb = Workbook()
    # --- Sheet 0: Field Guide ---
    guide = wb.active
    guide.title = "00_Field_Guide"
    guide["A1"] = "Article Evaluation Workbook — Revolution of Truth Pipeline"
    guide["A1"].font = Font(bold=True, size=14)
    rows = [
        ("Purpose", "One row per article (drv-00 … drv-06). Fill metrics here; run excel_to_site_json.py to emit verification + rigor JSON."),
        ("Outputs", "faiththruphysics-site-data/data-viz/verification-{slug}.json + rigor/{series}/{slug}.json"),
        ("Verified bar", "Maps to components/verification-bar.js → loadVerification(data)"),
        ("Domain pills", "01_Identity.domains_* columns + 03_Domains sheet (must sum to 100)"),
        ("Audit boxes", "07_Audit_Boxes sheet → got_right / overstated / got_wrong"),
        ("Claims layer", "04_Claims sheet → PhD proof layer (future)"),
        ("Physics process", "09_Physics_Process ← Station 02 / chi-evaluator double-check"),
        ("Narrative flow", "10_Narrative_Flow ← beginning/middle/end + 3 improve + 3 hurt"),
        ("Citations", "11_Recommended_Citations ← P04 paper recommender (when live)"),
        ("Station pipeline", "Optional: run 10-station batch; paste Station 10 summary into 02_Metrics"),
    ]
    for i, (k, v) in enumerate(rows, 3):
        guide.cell(i, 1, k).font = LABEL_FONT
        guide.cell(i, 2, v)

    # --- Sheet 1: Article Identity ---
    ident = wb.create_sheet("01_Article_Identity")
    ident_headers = [
        "series", "slug", "title", "subtitle", "doc_type", "reading_level_default",
        "proof_explorer_slug", "series_order", "prev_slug", "next_slug",
        "domain_1_name", "domain_1_pct", "domain_2_name", "domain_2_pct",
        "domain_3_name", "domain_3_pct", "domain_4_name", "domain_4_pct",
        "domain_5_name", "domain_5_pct", "domain_6_name", "domain_6_pct",
        "domain_7_name", "domain_7_pct", "domains_sum_check",
        "tags", "notes",
    ]
    header_row(ident, 1, ident_headers)
    sample = [
        "revolution-of-truth", "drv-00-the-argument", "The Argument in One Page", "",
        "series_article", "college", "drv-00-the-argument", 0, "", "drv-01-the-architecture",
        "Information Theory", 25, "Physics", 20, "Theology", 20, "Mathematics", 15,
        "Empirical Data", 10, "Consciousness", 5, "History/Culture", 5, 100,
        "coherence;lock-before-key;soteriological-limit", "",
    ]
    for col, val in enumerate(sample, 1):
        ident.cell(2, col, val)

    # --- Sheet 2: Verification Metrics (verified bar expand panel) ---
    metrics = wb.create_sheet("02_Verification_Metrics")
    metrics_headers = [
        "slug", "axioms_tested", "axioms_total", "axiom_ids_csv",
        "laws_active_csv", "chi_raw", "chi_normalized", "fruits_score",
        "iso_bridge_count", "iso_physics_processes", "iso_trinity_mappings", "iso_meq_variables",
        "claims_total", "claims_load_bearing", "claims_kill_conditions", "claims_contradictions",
        "qq0_posture", "qq1_identity", "qq2_domain", "qq3_claim", "qq4_support",
        "qq5_dependencies", "qq6_consequences", "qq7_kill_conditions", "inheritance_check",
        "lean_mode", "lean_theorem", "lean_status", "lean_file",
        "station10_summary", "evaluator", "eval_date",
    ]
    header_row(metrics, 1, metrics_headers)
    metrics_sample = [
        "drv-00-the-argument", 22, 188, "A1.1,A2.1,A5.1,BC4,BC6",
        "1,5,6,10", 7.4, 8.2, 7,
        4, 3, 2, 8,
        7, 5, 4, 0,
        "FILLED", "FILLED", "FILLED", "FILLED", "FILLED", "FILLED", "FILLED", "FILLED", "PASS",
        "B", "", "OPEN", "",
        "", "", "",
    ]
    for col, val in enumerate(metrics_sample, 1):
        metrics.cell(2, col, val)

    # --- Sheet 3: Domains (alternate long form) ---
    domains = wb.create_sheet("03_Domains")
    header_row(domains, 1, ["slug", "domain_name", "pct", "color_hex", "chip_on_page"])
    domain_rows = [
        ("drv-00-the-argument", "Information Theory", 25, "#a78bfa", "yes"),
        ("drv-00-the-argument", "Physics", 20, "#7cc7ff", "yes"),
        ("drv-00-the-argument", "Theology", 20, "#d4af37", "yes"),
        ("drv-00-the-argument", "Mathematics", 15, "#ff7d90", "yes"),
        ("drv-00-the-argument", "Empirical Data", 10, "#7fc77f", "yes"),
        ("drv-00-the-argument", "Consciousness", 5, "#a78bfa", "yes"),
        ("drv-00-the-argument", "History/Culture", 5, "#b8a088", "yes"),
    ]
    for r, row in enumerate(domain_rows, 2):
        for c, val in enumerate(row, 1):
            domains.cell(r, c, val)

    # --- Sheet 4: Claims ---
    claims = wb.create_sheet("04_Claims")
    header_row(claims, 1, [
        "slug", "claim_id", "claim_text", "load_bearing", "status",
        "parent_law", "operator_form", "derived_property", "kill_condition", "evidence",
    ])
    claim_samples = [
        ("drv-00-the-argument", "C1", "Coherence is the precondition for anything to exist, persist, or mean anything.", "yes", "STRUCTURALLY_MAPPED", "Law 6", "H(X|Y)→0", "Logos", "Find spiritual info term not derivable from Shannon at max mutual info", ""),
        ("drv-00-the-argument", "C2", "The Soteriological Limit closes the man-made math escape hatch.", "yes", "LEAN_DEFINED", "Law 1+5", "Gödel+Second Law", "Grace external", "Show finite-agent grounding without surrendering necessity/universality", ""),
    ]
    for r, row in enumerate(claim_samples, 2):
        for c, val in enumerate(row, 1):
            claims.cell(r, c, val)

    # --- Sheet 5: Axioms ---
    axioms = wb.create_sheet("05_Axioms")
    header_row(axioms, 1, ["slug", "axiom_id", "axiom_name", "tested", "role", "evidence_snippet"])
    axiom_samples = [
        ("drv-00-the-argument", "A1.1", "Existence", "yes", "core", "Something exists rather than nothing"),
        ("drv-00-the-argument", "A2.1", "Substrate Requirement", "yes", "core", "Information requires a substrate"),
        ("drv-00-the-argument", "BC4", "Three Observers Required", "yes", "boundary", "Triangulation without privileged frame"),
    ]
    for r, row in enumerate(axiom_samples, 2):
        for c, val in enumerate(row, 1):
            axioms.cell(r, c, val)

    # --- Sheet 6: Ten Laws ---
    laws = wb.create_sheet("06_Ten_Laws")
    header_row(laws, 1, ["slug", "law_num", "law_name", "active", "strength_1_10", "evidence"])
    law_names = [
        (1, "Gravity/Grace"), (2, "Mass/Meaning"), (3, "EM/Truth"), (4, "Strong/Love"),
        (5, "Thermo/Judgment"), (6, "Info/Logos"), (7, "Relativity/Relationship"),
        (8, "Quantum/Faith"), (9, "Weak/Sin-Conservation"), (10, "Coherence/Christ"),
    ]
    active = {1, 5, 6, 10}
    for i, (num, name) in enumerate(law_names, 2):
        laws.cell(i, 1, "drv-00-the-argument")
        laws.cell(i, 2, num)
        laws.cell(i, 3, name)
        laws.cell(i, 4, "yes" if num in active else "no")
        laws.cell(i, 5, 8 if num in active else "")
        laws.cell(i, 6, "")

    # --- Sheet 7: Audit boxes ---
    audit = wb.create_sheet("07_Audit_Boxes")
    header_row(audit, 1, ["slug", "bucket", "item_text", "source_station"])
    audit_samples = [
        ("drv-00-the-argument", "got_right", "Pre-linguistic numerosity and pre-socialized moral evaluation are load-bearing empirical anchors.", "05"),
        ("drv-00-the-argument", "got_right", "The lock-before-key method eliminates curve-fitting.", "04"),
        ("drv-00-the-argument", "overstated", "Probability band (1 in 1M to 1 in 100T) needs tighter independence documentation.", "05"),
        ("drv-00-the-argument", "got_wrong", "", ""),
    ]
    for r, row in enumerate(audit_samples, 2):
        for c, val in enumerate(row, 1):
            audit.cell(r, c, val)

    # --- Sheet 8: Kill conditions ---
    kills = wb.create_sheet("08_Kill_Conditions")
    header_row(kills, 1, ["slug", "kill_id", "condition", "test", "severity", "status"])
    kill_samples = [
        ("drv-00-the-argument", "K1", "Show man-made math grounding without surrendering necessity", "Philosophy of math audit", "fatal", "open"),
        ("drv-00-the-argument", "K2", "Find spiritual info term not derivable from Shannon framework", "Logos chain falsification", "fatal", "open"),
    ]
    for r, row in enumerate(kill_samples, 2):
        for c, val in enumerate(row, 1):
            kills.cell(r, c, val)

    # --- Sheet 9: Physics process (Station 02 / chi-evaluator cross-check) ---
    physics = wb.create_sheet("09_Physics_Process")
    header_row(physics, 1, [
        "slug", "maps_to_physics", "physics_domain", "specific_process",
        "equation_present", "isomorphism_detected", "isomorphism_type", "isomorphism_quality_1_10",
        "framework_alignment_total_0_100", "chi_coverage_pct", "source_engine", "notes",
    ])
    phys_sample = [
        "drv-00-the-argument", "yes", "information theory; thermodynamics; logic",
        "Shannon entropy; Gödel incompleteness; Second Law; Soteriological Limit",
        "yes", "yes", "structural", 8, 82, "100%", "api_call_02", "",
    ]
    for col, val in enumerate(phys_sample, 1):
        physics.cell(2, col, val)

    # --- Sheet 10: Narrative flow ---
    flow = wb.create_sheet("10_Narrative_Flow")
    header_row(flow, 1, [
        "slug", "flow_score_1_10", "beginning_summary", "middle_summary", "end_summary",
        "weakest_transition", "improve_1", "improve_2", "improve_3",
        "hurts_1", "hurts_2", "hurts_3", "source_station",
    ])
    flow_sample = [
        "drv-00-the-argument", 8,
        "Sets coherence as precondition; states three combined questions.",
        "Derives Soteriological Limit; empirical anchors; six-book arc.",
        "Open gaps named honestly; reading order offered.",
        "Book V isomorphism table may feel like a gear shift after Book IV test logic.",
        "Add a one-paragraph bridge before the isomorphism table.",
        "Tighten probability section with explicit independence assumptions.",
        "Move 'Open Gaps' earlier as a trust signal before the big claims.",
        "drv-00 tries to do overview + full arc + equations in one page — density spikes mid-article.",
        "Some inline math paragraphs run long without visual breaks (fixed by College enrichment).",
        "Probability band stated before method appendix is visible.",
        "04+08",
    ]
    for col, val in enumerate(flow_sample, 1):
        flow.cell(2, col, val)

    # --- Sheet 11: Recommended citations ---
    cites = wb.create_sheet("11_Recommended_Citations")
    header_row(cites, 1, [
        "slug", "rank", "paper_title", "authors", "year", "why_relevant", "citation_type", "source_engine",
    ])
    cite_samples = [
        ("drv-00-the-argument", 1, "On Computable Numbers", "Turing", 1936, "Finite-agent limits on computation", "foundational", "P04"),
        ("drv-00-the-argument", 2, "Über formal unentscheidbare Sätze", "Gödel", 1931, "Incompleteness underpins Soteriological Limit", "foundational", "P04"),
        ("drv-00-the-argument", 3, "Origins of number sense", "Lipton & Spelke", 2003, "Pre-linguistic numerosity evidence cited in paper", "empirical", "manual"),
    ]
    for r, row in enumerate(cite_samples, 2):
        for c, val in enumerate(row, 1):
            cites.cell(r, c, val)

    for ws in wb.worksheets:
        autosize(ws, min(ws.max_column, 12))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
