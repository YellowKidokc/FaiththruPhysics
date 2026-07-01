#!/usr/bin/env python3
"""
Convert Article_Evaluation_Workbook.xlsx → site JSON artifacts.

Outputs:
  faiththruphysics-site-data/data-viz/verification-{series}-{slug}.json
  faiththruphysics-site-data/rigor/{series}/{slug}.json
  faiththruphysics-site-data/claims/{series}/{slug}.json  (optional)

Usage:
  python excel_to_site_json.py
  python excel_to_site_json.py --workbook path/to/workbook.xlsx --slug drv-00-the-argument
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("pip install openpyxl") from exc

DATA = Path(r"D:\GitHub\faiththruphysics-site-data")
DEFAULT_WB = DATA / "APIs" / "revolution-of-truth-pipeline" / "templates" / "Article_Evaluation_Workbook.xlsx"


def sheet_rows(wb, name: str) -> list[dict]:
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for row in rows[1:]:
        if not any(row):
            continue
        out.append({headers[i]: row[i] for i in range(len(headers)) if headers[i]})
    return out


def parse_csv_nums(value) -> list[int]:
    if value is None or value == "":
        return []
    return [int(x.strip()) for x in str(value).split(",") if x.strip().isdigit()]


def build_verification(
    identity: dict,
    metrics: dict,
    domains: list[dict],
    laws: list[dict],
) -> dict:
    slug = str(metrics.get("slug") or identity.get("slug") or "").strip()
    series = str(identity.get("series") or "revolution-of-truth").strip()
    domain_map: dict[str, float] = {}
    for row in domains:
        if str(row.get("slug", "")).strip() != slug:
            continue
        name = str(row.get("domain_name", "")).strip()
        if not name:
            continue
        domain_map[name] = float(row.get("pct") or 0)
    if not domain_map:
        for i in range(1, 8):
            n = identity.get(f"domain_{i}_name")
            p = identity.get(f"domain_{i}_pct")
            if n and p not in (None, ""):
                domain_map[str(n).strip()] = float(p)

    laws_active = parse_csv_nums(metrics.get("laws_active_csv"))
    if not laws_active:
        laws_active = [
            int(row["law_num"])
            for row in laws
            if str(row.get("slug", "")).strip() == slug and str(row.get("active", "")).lower() == "yes"
        ]

    return {
        "schema_version": "1.0",
        "series": series,
        "slug": f"{series}/{slug}" if "/" not in slug else slug,
        "title": identity.get("title") or slug,
        "axioms": {
            "tested": int(metrics.get("axioms_tested") or 0),
            "total": int(metrics.get("axioms_total") or 188),
            "ids": [x.strip() for x in str(metrics.get("axiom_ids_csv") or "").split(",") if x.strip()],
        },
        "laws": {"active": sorted(set(laws_active))},
        "chi": {
            "raw": float(metrics.get("chi_raw") or 0),
            "normalized": float(metrics.get("chi_normalized") or 0),
        },
        "fruits": {"score": int(metrics.get("fruits_score") or 0)},
        "isomorphisms": {
            "count": int(metrics.get("iso_bridge_count") or 0),
            "physics_processes": int(metrics.get("iso_physics_processes") or 0),
            "trinity_mappings": int(metrics.get("iso_trinity_mappings") or 0),
            "meq_variables": int(metrics.get("iso_meq_variables") or 0),
        },
        "claims": {
            "total": int(metrics.get("claims_total") or 0),
            "load_bearing": int(metrics.get("claims_load_bearing") or 0),
            "kill_conditions": int(metrics.get("claims_kill_conditions") or 0),
            "contradictions": int(metrics.get("claims_contradictions") or 0),
        },
        "domains": domain_map,
        "framework": {
            "qq0_posture": metrics.get("qq0_posture") or "",
            "qq1_identity": metrics.get("qq1_identity") or "",
            "qq2_domain": metrics.get("qq2_domain") or "",
            "qq3_claim": metrics.get("qq3_claim") or "",
            "qq4_support": metrics.get("qq4_support") or "",
            "qq5_dependencies": metrics.get("qq5_dependencies") or "",
            "qq6_consequences": metrics.get("qq6_consequences") or "",
            "qq7_kill_conditions": metrics.get("qq7_kill_conditions") or "",
            "inheritance_check": metrics.get("inheritance_check") or "",
            "lean": {
                "mode": metrics.get("lean_mode") or "",
                "theorem": metrics.get("lean_theorem") or "",
                "status": metrics.get("lean_status") or "",
                "lean_file": metrics.get("lean_file") or "",
            },
        },
        "meta": {
            "evaluator": metrics.get("evaluator") or "",
            "eval_date": str(metrics.get("eval_date") or ""),
            "station10_summary": metrics.get("station10_summary") or "",
        },
    }


def build_rigor(slug: str, audit_rows: list[dict]) -> dict:
    got_right, overstated, got_wrong = [], [], []
    for row in audit_rows:
        if str(row.get("slug", "")).strip() != slug:
            continue
        bucket = str(row.get("bucket", "")).strip().lower()
        text = str(row.get("item_text") or "").strip()
        if not text or text.lower() == "none":
            continue
        if bucket == "got_right":
            got_right.append(text)
        elif bucket == "overstated":
            overstated.append(text)
        elif bucket == "got_wrong":
            got_wrong.append(text)
    return {"got_right": got_right, "overstated": overstated, "got_wrong": got_wrong}


def build_claims(slug: str, claim_rows: list[dict]) -> list[dict]:
    out = []
    for row in claim_rows:
        if str(row.get("slug", "")).strip() != slug:
            continue
        text = str(row.get("claim_text", "")).strip()
        if not text:
            continue
        out.append({
            "id": row.get("claim_id") or "",
            "text": text,
            "load_bearing": str(row.get("load_bearing", "")).lower() == "yes",
            "status": row.get("status") or "",
            "parent_law": row.get("parent_law") or "",
            "operator_form": row.get("operator_form") or "",
            "derived_property": row.get("derived_property") or "",
            "kill_condition": row.get("kill_condition") or "",
            "evidence": row.get("evidence") or "",
        })
    return out


def build_editorial(
    slug: str,
    physics_rows: list[dict],
    flow_rows: list[dict],
    cite_rows: list[dict],
) -> dict:
    physics = next((r for r in physics_rows if str(r.get("slug", "")).strip() == slug), {})
    flow = next((r for r in flow_rows if str(r.get("slug", "")).strip() == slug), {})
    cites = [
        {
            "rank": int(row.get("rank") or 0),
            "title": row.get("paper_title") or "",
            "authors": row.get("authors") or "",
            "year": row.get("year") or "",
            "why_relevant": row.get("why_relevant") or "",
            "citation_type": row.get("citation_type") or "",
            "source": row.get("source_engine") or "",
        }
        for row in cite_rows
        if str(row.get("slug", "")).strip() == slug and row.get("paper_title")
    ]
    cites.sort(key=lambda c: c["rank"] or 99)
    return {
        "physics_process": {
            "maps_to_physics": str(physics.get("maps_to_physics", "")).lower() in ("yes", "true", "1"),
            "physics_domain": physics.get("physics_domain") or "",
            "specific_process": physics.get("specific_process") or "",
            "equation_present": str(physics.get("equation_present", "")).lower() in ("yes", "true", "1"),
            "isomorphism": {
                "detected": str(physics.get("isomorphism_detected", "")).lower() in ("yes", "true", "1"),
                "type": physics.get("isomorphism_type") or "",
                "quality": int(physics.get("isomorphism_quality_1_10") or 0),
            },
            "framework_alignment": int(physics.get("framework_alignment_total_0_100") or 0),
            "chi_coverage_pct": physics.get("chi_coverage_pct") or "",
            "source": physics.get("source_engine") or "",
        },
        "narrative_flow": {
            "score": int(flow.get("flow_score_1_10") or 0),
            "arc": {
                "beginning": flow.get("beginning_summary") or "",
                "middle": flow.get("middle_summary") or "",
                "end": flow.get("end_summary") or "",
            },
            "weakest_transition": flow.get("weakest_transition") or "",
            "improve": [flow.get(f"improve_{i}") or "" for i in range(1, 4)],
            "hurts": [flow.get(f"hurts_{i}") or "" for i in range(1, 4)],
            "source": flow.get("source_station") or "",
        },
        "recommended_citations": cites,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WB)
    parser.add_argument("--slug", help="Only export one slug")
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    if not args.workbook.is_file():
        print(f"Missing workbook: {args.workbook}")
        print("Run: python build_article_evaluation_workbook.py")
        return 1

    wb = load_workbook(args.workbook, data_only=True)
    identity_rows = sheet_rows(wb, "01_Article_Identity")
    metrics_rows = sheet_rows(wb, "02_Verification_Metrics")
    domain_rows = sheet_rows(wb, "03_Domains")
    claim_rows = sheet_rows(wb, "04_Claims")
    law_rows = sheet_rows(wb, "06_Ten_Laws")
    audit_rows = sheet_rows(wb, "07_Audit_Boxes")
    physics_rows = sheet_rows(wb, "09_Physics_Process")
    flow_rows = sheet_rows(wb, "10_Narrative_Flow")
    cite_rows = sheet_rows(wb, "11_Recommended_Citations")

    identity_by_slug = {str(r.get("slug", "")).strip(): r for r in identity_rows if r.get("slug")}
    slugs = [args.slug] if args.slug else list(identity_by_slug.keys())

    for slug in slugs:
        identity = identity_by_slug.get(slug, {"slug": slug, "series": "revolution-of-truth"})
        metrics = next((r for r in metrics_rows if str(r.get("slug", "")).strip() == slug), {"slug": slug})
        series = str(identity.get("series") or "revolution-of-truth").strip()
        verification = build_verification(identity, metrics, domain_rows, law_rows)
        rigor = build_rigor(slug, audit_rows)
        claims = build_claims(slug, claim_rows)
        editorial = build_editorial(slug, physics_rows, flow_rows, cite_rows)

        safe = slug.replace("/", "-")
        viz_path = DATA / "data-viz" / f"verification-{series}-{safe}.json"
        rigor_path = DATA / "rigor" / series / f"{slug}.json"
        claims_path = DATA / "claims" / series / f"{slug}.json"
        editorial_path = DATA / "editorial" / series / f"{slug}.json"

        print(f"\n{series}/{slug}")
        print(f"  verification -> {viz_path}")
        print(f"  rigor        -> {rigor_path}")
        print(f"  claims       -> {claims_path} ({len(claims)} items)")
        print(f"  editorial    -> {editorial_path}")

        if args.apply:
            viz_path.parent.mkdir(parents=True, exist_ok=True)
            rigor_path.parent.mkdir(parents=True, exist_ok=True)
            claims_path.parent.mkdir(parents=True, exist_ok=True)
            editorial_path.parent.mkdir(parents=True, exist_ok=True)
            viz_path.write_text(json.dumps(verification, indent=2), encoding="utf-8")
            rigor_path.write_text(json.dumps(rigor, indent=2), encoding="utf-8")
            claims_path.write_text(json.dumps(claims, indent=2), encoding="utf-8")
            editorial_path.write_text(json.dumps(editorial, indent=2), encoding="utf-8")

    if not args.apply:
        print("\nRe-run with --apply to write JSON files.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
