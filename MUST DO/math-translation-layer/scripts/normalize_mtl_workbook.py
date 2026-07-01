#!/usr/bin/env python3
"""Normalize an MTL workbook into JSON/CSV plus needs-review outputs.

This is intentionally conservative:
- primary sheet can define rich records
- fallback sheets can fill lighter fields
- ambiguous rows stay in needs-review outputs
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from pathlib import Path

from openpyxl import load_workbook


def normalize_latex(value: str) -> str:
    text = (value or "").strip()
    text = re.sub(r"^\$+", "", text)
    text = re.sub(r"\$+$", "", text)
    text = re.sub(r"^\\\(|\\\)$", "", text)
    text = re.sub(r"^\\\[|\\\]$", "", text)
    return text.strip()


def latex_hash(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def write_csv(rows: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames: list[str] = []
    for row in rows:
        for key in row:
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def sheet_rows(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    out = []
    for row in rows[1:]:
        record = {}
        for key, value in zip(headers, row):
            if key:
                record[key] = value
        out.append(record)
    return out


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Normalize an MTL workbook.")
    parser.add_argument("--workbook", required=True, help="Path to workbook")
    parser.add_argument("--output-dir", required=True, help="Output folder")
    parser.add_argument("--primary-sheet", default="Consciousness_Equations", help="Primary rich sheet")
    parser.add_argument("--fallback-sheet", default="Existing_458_Translations", help="Fallback translation sheet")
    parser.add_argument("--core-sheet", default="Core_16_Equations", help="Core equation sheet")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook_path = Path(args.workbook)
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(workbook_path, read_only=True, data_only=False)

    primary = sheet_rows(wb[args.primary_sheet])
    fallback = sheet_rows(wb[args.fallback_sheet]) if args.fallback_sheet in wb.sheetnames else []
    core = sheet_rows(wb[args.core_sheet]) if args.core_sheet in wb.sheetnames else []

    fallback_by_hash: dict[str, dict] = {}
    for row in fallback:
        latex = normalize_latex(str(row.get("latex") or ""))
        if not latex:
            continue
        fallback_by_hash[latex_hash(latex)] = row

    core_by_hash: dict[str, dict] = {}
    for row in core:
        latex = normalize_latex(str(row.get("latex") or ""))
        if not latex:
            continue
        core_by_hash[latex_hash(latex)] = row

    normalized = []
    needs_review = []

    for row in primary:
        raw = str(row.get("Equation_Raw") or "")
        latex = normalize_latex(raw)
        if not latex:
            continue
        digest = latex_hash(latex)
        fallback_row = fallback_by_hash.get(digest, {})
        core_row = core_by_hash.get(digest, {})

        record = {
            "eq_id": str(row.get("ID") or "").strip(),
            "type": str(row.get("Type") or "").strip(),
            "latex": latex,
            "latex_hash": digest,
            "source_file": str(row.get("Source_File") or "").strip(),
            "context": str(row.get("Context") or "").strip(),
            "easy": str(row.get("English_Translation") or "").strip() or str(core_row.get("plain_english") or "").strip(),
            "standard": str(row.get("Term_By_Term") or "").strip() or str(fallback_row.get("medium_form") or "").strip(),
            "academic": str(row.get("Physics_Side") or "").strip(),
            "theology_side": str(row.get("Theology_Side") or "").strip(),
            "shared_structure": str(row.get("Shared_Structure") or "").strip(),
            "conceptual_meaning": str(fallback_row.get("conceptual_meaning") or core_row.get("conceptual_meaning") or "").strip(),
            "audio_safe": str(fallback_row.get("tts_audio") or core_row.get("tts_audio") or "").strip(),
            "difficulty": str(row.get("Difficulty") or "").strip(),
            "law_reference": str(row.get("Law_Reference") or "").strip(),
            "paper_ref": str(fallback_row.get("paper_ref") or core_row.get("paper_ref") or "").strip(),
            "needs_review": False,
        }

        rich_fields = [
            record["easy"],
            record["standard"],
            record["academic"],
            record["theology_side"],
            record["shared_structure"],
        ]
        filled_rich = sum(1 for item in rich_fields if item)

        if filled_rich < 2:
            record["needs_review"] = True
            needs_review.append(record)
        normalized.append(record)

    summary = {
        "workbook": str(workbook_path),
        "primary_sheet_rows": len(primary),
        "fallback_sheet_rows": len(fallback),
        "core_sheet_rows": len(core),
        "normalized_rows": len(normalized),
        "needs_review_rows": len(needs_review),
        "ready_rows": len([row for row in normalized if not row["needs_review"]]),
    }

    (output_dir / "normalized_mtl_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    (output_dir / "normalized_mtl_records.json").write_text(json.dumps(normalized, indent=2, ensure_ascii=False), encoding="utf-8")
    (output_dir / "normalized_mtl_needs_review.json").write_text(json.dumps(needs_review, indent=2, ensure_ascii=False), encoding="utf-8")
    write_csv(normalized, output_dir / "normalized_mtl_records.csv")
    write_csv(needs_review, output_dir / "normalized_mtl_needs_review.csv")


if __name__ == "__main__":
    main()
